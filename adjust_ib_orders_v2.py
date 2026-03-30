#!/usr/bin/env python3
"""
Place new limit orders and continuously adjust their prices via Interactive Brokers API.

Reads an Excel file specifying contracts to order, then fetches market prices, places
orders (untransmitted), asks for confirmation, transmits them, and runs an adjustment loop.

Excel format: Sheet named "inputs" with columns:
  - CONID: IB contract ID (required)
  - ACTION: BUY or SELL (required)
  - QTY: order quantity (required)
  - [blank column]
  - TICK_BASED: if 1, delta values are in ticks (multiplied by contract minTick); blank = raw price
  - DELTA1, TIME1, PRICE1: first tier (delta, minutes, price threshold)
  - DELTA2, TIME2, PRICE2: second tier (delta, minutes, price threshold)
  - DELTA3, TIME3: final tier (delta, minutes, no threshold)

Tier logic (BUY orders raise price, SELL orders lower price):
  - BUY: use tier 1 while price < PRICE1, tier 2 while price < PRICE2, then tier 3
  - SELL: use tier 1 while price > PRICE1, tier 2 while price > PRICE2, then tier 3

Usage:
    python3 adjust_ib_orders_v2.py [excel_file] [-y]

Prerequisites:
  1. pip install ibapi openpyxl
  2. TWS or IB Gateway running on localhost
"""

import argparse
import sys
import threading
import time
import warnings
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Optional

# Suppress openpyxl warnings about Excel styling
warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')

from openpyxl import load_workbook

from ibapi.client import EClient
from ibapi.wrapper import EWrapper
from ibapi.contract import Contract
from ibapi.order import Order
from ibapi.common import TickerId


# =============================================================================
# CONFIGURATION
# =============================================================================

SCRIPT_DIR = Path(__file__).resolve().parent
DEFAULT_EXCEL_PATH = SCRIPT_DIR / "adjust_ib_inputs2.xlsx"

# IB Connection settings
DEFAULT_HOST = "127.0.0.1"
DEFAULT_PORT = 7497  # 7497=TWS paper, 7496=TWS live, 4002=Gateway
DEFAULT_CLIENT_ID = 7  # Client ID 7: adjust_ib_orders_v2.py (Trading range 1-9)

# Minimum tick size for price rounding
DEFAULT_TICK_SIZE = 0.01

# Market data fetch timeout (seconds)
MARKET_DATA_TIMEOUT = 10


# =============================================================================
# DATA CLASSES
# =============================================================================

@dataclass
class AdjustmentConfig:
    """Configuration for a single order to place and adjust."""
    conid: int                    # IB contract ID
    action: str                   # BUY or SELL
    qty: float                    # Order quantity
    tiers: list                   # [{delta, interval_sec, threshold}, ...]
    tick_based: bool = False      # If True, delta values are in ticks (multiplied by minTick)
    # Runtime state
    order_id: Optional[int] = None  # Assigned after order placement
    last_adjusted: float = 0.0
    adjustment_count: int = 0
    current_tier: int = 0


@dataclass
class OrderInfo:
    """Information about an open order."""
    order_id: int
    contract: Contract
    order: Order
    action: str
    limit_price: float
    quantity: float
    symbol: str
    sec_type: str


# =============================================================================
# EXCEL READING
# =============================================================================

def read_adjust_configs(excel_path):
    """Read the adjustment Excel file and return a list of AdjustmentConfig objects."""
    wb = load_workbook(excel_path, data_only=True)
    ws = wb["inputs"]

    # Read header row — strip whitespace, skip blank/None headers
    headers = []
    for cell in ws[1]:
        val = cell.value
        if val is not None:
            stripped = str(val).strip()
            headers.append(stripped if stripped else None)
        else:
            headers.append(None)

    configs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Skip empty rows
        if not row or all(cell is None for cell in row):
            continue

        # Build dict skipping blank/None header keys
        row_dict = {}
        for header, val in zip(headers, row):
            if header:
                row_dict[header] = val

        conid_val = row_dict.get("CONID")
        action_val = row_dict.get("ACTION")
        qty_val = row_dict.get("QTY")

        # Tick-based flag
        tick_based_val = row_dict.get("TICK_BASED")
        tick_based = tick_based_val is not None and str(tick_based_val).strip() == "1"

        # Tier values
        delta1 = row_dict.get("DELTA1")
        time1 = row_dict.get("TIME1")
        price1 = row_dict.get("PRICE1")
        delta2 = row_dict.get("DELTA2")
        time2 = row_dict.get("TIME2")
        price2 = row_dict.get("PRICE2")
        delta3 = row_dict.get("DELTA3")
        time3 = row_dict.get("TIME3")

        # Skip rows with missing required fields
        if conid_val is None or action_val is None or qty_val is None:
            continue

        conid = int(conid_val)
        if conid < 0:
            continue
        action = str(action_val).strip().upper()
        qty = float(qty_val)

        if action not in ("BUY", "SELL"):
            print(f"  WARNING: skipping row with invalid ACTION '{action}'")
            continue

        if qty <= 0:
            print(f"  WARNING: skipping row with non-positive QTY {qty}")
            continue

        # Build tiers dynamically
        tiers = []

        # Tier 1 (required, PRICE1 optional)
        if delta1 is None or time1 is None:
            print(f"  WARNING: skipping row missing DELTA1/TIME1")
            continue
        delta1 = float(delta1)
        time1 = float(time1)
        if delta1 <= 0 or time1 <= 0:
            print(f"  WARNING: skipping row with non-positive DELTA1 or TIME1")
            continue
        threshold1 = float(price1) if price1 is not None else None
        tiers.append({"delta": delta1, "interval_sec": time1 * 60, "threshold": threshold1})

        # Tier 2 (optional, but all fields required if adding)
        if delta2 is not None and time2 is not None and price2 is not None:
            delta2 = float(delta2)
            time2 = float(time2)
            price2 = float(price2)
            if delta2 <= 0 or time2 <= 0:
                print(f"  WARNING: skipping row with non-positive DELTA2 or TIME2")
                continue
            tiers.append({"delta": delta2, "interval_sec": time2 * 60, "threshold": price2})
        elif any(val is not None for val in (delta2, time2, price2)):
            print(f"  WARNING: skipping row with partial tier 2 (all DELTA2/TIME2/PRICE2 required if any provided)")
            continue

        # Tier 3 (optional, no PRICE3)
        if delta3 is not None and time3 is not None:
            delta3 = float(delta3)
            time3 = float(time3)
            if delta3 <= 0 or time3 <= 0:
                print(f"  WARNING: skipping row with non-positive DELTA3 or TIME3")
                continue
            tiers.append({"delta": delta3, "interval_sec": time3 * 60, "threshold": None})
        elif any(val is not None for val in (delta3, time3)):
            print(f"  WARNING: skipping row with partial tier 3 (both DELTA3/TIME3 required if any provided)")
            continue

        # Consistency checks
        if not tiers:
            continue

        # No non-final tier should have None threshold
        if any(tier["threshold"] is None for tier in tiers[:-1]):
            if len(tiers) > 1 or tiers[0]["threshold"] is not None:
                print(f"  WARNING: skipping row with non-final tier missing threshold")
                continue

        # If first tier has None threshold but there are higher tiers, doesn't make sense
        if tiers[0]["threshold"] is None and len(tiers) > 1:
            print(f"  WARNING: skipping row (first tier has no threshold, higher tiers unreachable)")
            continue

        # If last tier has a threshold (i.e., no explicit final tier), append a final one using same delta/interval
        if tiers[-1]["threshold"] is not None:
            last_tier = tiers[-1]
            tiers.append({"delta": last_tier["delta"], "interval_sec": last_tier["interval_sec"], "threshold": None})

        configs.append(AdjustmentConfig(
            conid=conid,
            action=action,
            qty=qty,
            tiers=tiers,
            tick_based=tick_based,
        ))

    wb.close()
    return configs


def get_current_tier(cfg: AdjustmentConfig, current_price: float):
    """Return (tier_index, delta, interval_sec) based on price thresholds."""
    tiers = cfg.tiers
    action = cfg.action

    for i, tier in enumerate(tiers[:-1]):  # Check all but last tier
        threshold = tier["threshold"]
        if action == "BUY" and current_price < threshold:
            return i, tier["delta"], tier["interval_sec"]
        elif action == "SELL" and current_price > threshold:
            return i, tier["delta"], tier["interval_sec"]

    # Final tier (no threshold)
    return len(tiers) - 1, tiers[-1]["delta"], tiers[-1]["interval_sec"]


def get_current_interval(cfg: AdjustmentConfig):
    """Get the interval for the current tier."""
    tier_idx = cfg.current_tier
    return cfg.tiers[tier_idx]["interval_sec"]


# =============================================================================
# IB API WRAPPER AND CLIENT
# =============================================================================

class IBOrderAdjuster(EWrapper, EClient):
    """IB API client for placing and adjusting order prices."""

    def __init__(self):
        EClient.__init__(self, self)

        # Threading
        self.lock = threading.Lock()
        self.connected_event = threading.Event()
        self.orders_received_event = threading.Event()

        # Order storage
        self.open_orders: dict[int, OrderInfo] = {}
        self.next_order_id: int = 0

        # Tick size cache: conId -> minTick
        self.tick_sizes: dict[int, float] = {}
        self._tick_size_events: dict[int, threading.Event] = {}

        # Resolved contracts cache: conId -> Contract (full details from IB)
        self.resolved_contracts: dict[int, Contract] = {}

        # Market data: reqId -> {bid, ask}
        self.market_data: dict[int, dict] = {}
        self._market_data_events: dict[int, threading.Event] = {}

        # Transmit tracking: order IDs waiting for user to transmit in TWS
        self.awaiting_transmit: set[int] = set()
        self.all_transmitted_event = threading.Event()

        # State
        self.running: bool = True

    # -------------------------------------------------------------------------
    # Connection Callbacks
    # -------------------------------------------------------------------------

    def connectAck(self):
        """Called when connection is acknowledged."""
        pass

    def nextValidId(self, orderId: int):
        """Receives next valid order ID - signals connection is ready."""
        self.next_order_id = orderId
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Connected. Next valid order ID: {orderId}")
        self.connected_event.set()

    def error(self, reqId: TickerId, errorCode: int, errorString: str,
              advancedOrderRejectJson: str = ""):
        """Handle errors from IB."""
        # Filter out non-critical messages
        if errorCode in [2104, 2106, 2158, 2119]:  # Market data farm messages
            pass  # Silently ignore
        elif errorCode == 10167:  # Delayed market data
            pass  # Silently ignore
        elif errorCode == 202:  # Order cancelled
            pass  # Expected when orders fill or get cancelled
        else:
            print(f"[ERROR] ReqId: {reqId}, Code: {errorCode}, Msg: {errorString}")

    def connectionClosed(self):
        """Called when connection is closed."""
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Connection closed")
        self.running = False

    # -------------------------------------------------------------------------
    # Order Callbacks
    # -------------------------------------------------------------------------

    def openOrder(self, orderId: int, contract: Contract, order: Order,
                  orderState):
        """Receive open order information."""
        with self.lock:
            action = order.action  # BUY or SELL

            if contract.secType == "BAG":
                symbol = contract.symbol + " combo"
            else:
                symbol = contract.symbol

            # Only track limit orders
            if order.orderType == "LMT":
                self.open_orders[orderId] = OrderInfo(
                    order_id=orderId,
                    contract=contract,
                    order=order,
                    action=action,
                    limit_price=order.lmtPrice,
                    quantity=order.totalQuantity,
                    symbol=symbol,
                    sec_type=contract.secType,
                )

    def openOrderEnd(self):
        """Called when all open orders have been received."""
        self.orders_received_event.set()

    def orderStatus(self, orderId: int, status: str, filled,
                    remaining, avgFillPrice: float, permId: int,
                    parentId: int, lastFillPrice: float, clientId: int,
                    whyHeld: str, mktCapPrice: float):
        """Receive order status updates."""
        if status == "Submitted":
            with self.lock:
                if orderId in self.awaiting_transmit:
                    self.awaiting_transmit.discard(orderId)
                    print(f"[{datetime.now().strftime('%H:%M:%S')}] Order #{orderId} transmitted → active")
                    if not self.awaiting_transmit:
                        self.all_transmitted_event.set()

        # Remove filled/cancelled orders from tracking
        if status in ("Filled", "Cancelled", "Inactive"):
            with self.lock:
                if orderId in self.open_orders:
                    del self.open_orders[orderId]

    # -------------------------------------------------------------------------
    # Market Data Callbacks
    # -------------------------------------------------------------------------

    def tickPrice(self, reqId: TickerId, tickType: int, price: float, attrib):
        """Receive price tick data."""
        with self.lock:
            if reqId not in self.market_data:
                self.market_data[reqId] = {}
            # tickType 1 = BID, tickType 2 = ASK
            if tickType == 1 and price > 0:
                self.market_data[reqId]["bid"] = price
            elif tickType == 2 and price > 0:
                self.market_data[reqId]["ask"] = price

            # Signal if we have both bid and ask
            data = self.market_data[reqId]
            if "bid" in data and "ask" in data:
                if reqId in self._market_data_events:
                    self._market_data_events[reqId].set()

    # -------------------------------------------------------------------------
    # Contract Details Callbacks
    # -------------------------------------------------------------------------

    def contractDetails(self, reqId: int, contractDetails):
        """Receive contract details — extract minTick and cache full contract."""
        con_id = contractDetails.contract.conId
        min_tick = contractDetails.minTick
        with self.lock:
            self.tick_sizes[con_id] = min_tick
            self.resolved_contracts[con_id] = contractDetails.contract

    def contractDetailsEnd(self, reqId: int):
        """Called when contract details request is complete."""
        with self.lock:
            if reqId in self._tick_size_events:
                self._tick_size_events[reqId].set()

    def resolve_contract(self, conid: int) -> Contract:
        """Return a fully resolved Contract from IB for any instrument type."""
        with self.lock:
            if conid in self.resolved_contracts:
                return self.resolved_contracts[conid]

        stub = Contract()
        stub.conId = conid  # IB resolves secType/exchange/currency from conId alone

        req_id = self.next_order_id
        self.next_order_id += 1
        event = threading.Event()
        with self.lock:
            self._tick_size_events[req_id] = event

        self.reqContractDetails(req_id, stub)
        event.wait(timeout=10)

        with self.lock:
            self._tick_size_events.pop(req_id, None)
            contract = self.resolved_contracts.get(conid)

        if contract is None:
            raise RuntimeError(f"Could not resolve contract for CONID {conid}")
        return contract

    def get_tick_size(self, contract: Contract) -> float:
        """Return the minTick for a contract. Uses cache; resolves via IB if needed."""
        con_id = contract.conId
        with self.lock:
            if con_id in self.tick_sizes:
                return self.tick_sizes[con_id]
        # resolve_contract populates tick_sizes as a side effect
        self.resolve_contract(con_id)
        with self.lock:
            return self.tick_sizes.get(con_id, DEFAULT_TICK_SIZE)

    # -------------------------------------------------------------------------
    # Market Data Methods
    # -------------------------------------------------------------------------

    def get_initial_price(self, conid: int, action: str) -> float:
        """Fetch market bid/ask for a contract and return the appropriate side price.

        Returns bid price for SELL orders, ask price for BUY orders.
        Raises RuntimeError if prices not received within timeout.
        """
        contract = self.resolve_contract(conid)

        req_id = self.next_order_id
        self.next_order_id += 1

        event = threading.Event()
        with self.lock:
            self.market_data[req_id] = {}
            self._market_data_events[req_id] = event

        # Request snapshot market data (generic_tick_list="" for default ticks)
        self.reqMktData(req_id, contract, "", True, False, [])

        # Wait for both bid and ask
        received = event.wait(timeout=MARKET_DATA_TIMEOUT)

        # Cancel market data subscription
        self.cancelMktData(req_id)

        with self.lock:
            self._market_data_events.pop(req_id, None)
            data = self.market_data.pop(req_id, {})

        if not received or ("bid" not in data and "ask" not in data):
            raise RuntimeError(
                f"No market data received for CONID {conid} within {MARKET_DATA_TIMEOUT}s"
            )

        bid = data.get("bid")
        ask = data.get("ask")

        if action == "SELL":
            if bid is not None:
                return bid
            if ask is not None:
                return ask
        else:  # BUY
            if ask is not None:
                return ask
            if bid is not None:
                return bid

        raise RuntimeError(f"No valid price received for CONID {conid}")

    # -------------------------------------------------------------------------
    # Order Placement Methods
    # -------------------------------------------------------------------------

    def place_order_for_config(self, cfg: AdjustmentConfig, initial_price: float):
        """Build contract and limit order, place with transmit=False, store order_id in cfg."""
        contract = self.resolve_contract(cfg.conid)  # already cached from get_initial_price

        order = Order()
        order.action = cfg.action
        order.orderType = "LMT"
        order.totalQuantity = cfg.qty
        order.lmtPrice = initial_price
        order.transmit = False

        order_id = self.next_order_id
        self.next_order_id += 1

        with self.lock:
            self.awaiting_transmit.add(order_id)
        self.placeOrder(order_id, contract, order)
        cfg.order_id = order_id

    # -------------------------------------------------------------------------
    # Helper Methods
    # -------------------------------------------------------------------------

    def _round_price(self, price: float, tick_size: float = DEFAULT_TICK_SIZE) -> float:
        """Round price to tick size."""
        tick = Decimal(str(tick_size))
        rounded = Decimal(str(price)).quantize(tick, rounding=ROUND_HALF_UP)
        return float(rounded)

    def request_open_orders(self):
        """Request all open orders and wait for response."""
        with self.lock:
            self.open_orders.clear()
        self.orders_received_event.clear()
        self.reqAllOpenOrders()
        self.orders_received_event.wait(timeout=5)

    def find_matching_order(self, cfg: AdjustmentConfig) -> Optional[OrderInfo]:
        """Find an open order matching the config by order ID."""
        with self.lock:
            return self.open_orders.get(cfg.order_id)

    def modify_order(self, order_info: OrderInfo, new_price: float) -> bool:
        """Modify an order's limit price. Returns True on success."""
        tick_size = self.get_tick_size(order_info.contract)
        new_price = self._round_price(new_price, tick_size)

        new_order = Order()
        new_order.action = order_info.order.action
        new_order.orderType = "LMT"
        new_order.lmtPrice = new_price
        new_order.totalQuantity = order_info.order.totalQuantity
        new_order.tif = order_info.order.tif
        new_order.outsideRth = order_info.order.outsideRth

        try:
            self.placeOrder(order_info.order_id, order_info.contract, new_order)
            order_info.order = new_order
            order_info.limit_price = new_price
            return True
        except Exception as e:
            print(f"  FAIL modify order #{order_info.order_id}: {e}")
            return False


# =============================================================================
# MAIN LOOP
# =============================================================================

def timestamp():
    """Return current local time as [HH:MM:SS] string."""
    return datetime.now().strftime("[%H:%M:%S]")


def run_adjustment_loop(adjuster: IBOrderAdjuster, configs: list[AdjustmentConfig]):
    """Main loop: adjust orders on schedule until all are filled or user interrupts."""
    now = time.time()
    for cfg in configs:
        cfg.last_adjusted = now

    total_adjustments = 0
    adjustment_log = []

    print()
    print(f"Monitoring {len(configs)} order(s). Press Ctrl+C to stop.")
    print()

    try:
        while configs and adjuster.running:
            now = time.time()

            next_time = min(
                cfg.last_adjusted + get_current_interval(cfg) for cfg in configs
            )
            wait = next_time - now
            if wait > 0:
                time.sleep(min(wait, 1.0))
                continue

            now = time.time()

            due = [cfg for cfg in configs
                   if now >= cfg.last_adjusted + get_current_interval(cfg)]

            if not due:
                continue

            adjuster.request_open_orders()

            to_remove = []

            for cfg in due:
                order_info = adjuster.find_matching_order(cfg)

                if order_info is None:
                    remaining = len(configs) - len(to_remove) - 1
                    print(f"{timestamp()} FILLED/GONE: Order #{cfg.order_id} {cfg.action} "
                          f"— removing from rotation ({remaining} remaining)")
                    to_remove.append(cfg)
                    continue

                current_price = order_info.limit_price
                tier_idx, delta, interval_sec = get_current_tier(cfg, current_price)

                if cfg.tick_based:
                    tick_size = adjuster.get_tick_size(order_info.contract)
                    delta = delta * tick_size

                old_tier = cfg.current_tier
                cfg.current_tier = tier_idx

                if cfg.action == "BUY":
                    new_price = current_price + delta
                else:
                    new_price = current_price - delta

                if new_price < 0:
                    new_price = 0.01

                old_price = current_price
                success = adjuster.modify_order(order_info, new_price)

                cfg.last_adjusted = time.time()

                if success:
                    cfg.adjustment_count += 1
                    total_adjustments += 1

                    tier_label = f"T{tier_idx + 1}"
                    tier_change = ""
                    if old_tier != tier_idx:
                        tier_change = f" [was T{old_tier + 1}]"

                    new_price_rounded = adjuster._round_price(new_price)
                    print(f"{timestamp()} {tier_label}{tier_change} Order #{cfg.order_id} {cfg.action}: "
                          f"${old_price:.2f} -> ${new_price_rounded:.2f} "
                          f"(#{cfg.adjustment_count})")
                    adjustment_log.append({
                        "time": datetime.now().strftime("%H:%M:%S"),
                        "order_id": cfg.order_id,
                        "action": cfg.action,
                        "old_price": old_price,
                        "new_price": new_price_rounded,
                        "tier": tier_idx + 1,
                    })

            for cfg in to_remove:
                configs.remove(cfg)

            if not configs:
                print()
                print("All orders filled or gone. Exiting.")

    except KeyboardInterrupt:
        print()
        print()
        print("Interrupted by user.")

    # Summary
    print()
    print(f"Summary: {total_adjustments} adjustment(s) made")
    if adjustment_log:
        print()
        print(f"  {'TIME':<10} {'TIER':<4} {'ACTION':<6} {'OLD':>8} {'NEW':>8}  ORDER_ID")
        print(f"  {'-'*10} {'-'*4} {'-'*6} {'-'*8} {'-'*8}  {'-'*10}")
        for entry in adjustment_log:
            print(f"  {entry['time']:<10} T{entry['tier']:<3} {entry['action']:<6} "
                  f"${entry['old_price']:>7.2f} ${entry['new_price']:>7.2f}  "
                  f"#{entry['order_id']}")
    print()


# =============================================================================
# MAIN
# =============================================================================

def parse_args():
    parser = argparse.ArgumentParser(
        description="Place new limit orders and continuously adjust their prices via IB.")
    parser.add_argument("excel_file", nargs="?", default=None,
                        help=f"Path to adjustment Excel file (default: {DEFAULT_EXCEL_PATH.name})")
    parser.add_argument("--yes", "-y", action="store_true",
                        help="Skip confirmation prompts")
    parser.add_argument("--port", "-p", type=int, default=DEFAULT_PORT,
                        help=f"TWS/Gateway port (default: {DEFAULT_PORT})")
    parser.add_argument("--client-id", "-c", type=int, default=DEFAULT_CLIENT_ID,
                        help=f"Client ID (default: {DEFAULT_CLIENT_ID})")
    return parser.parse_args()


def main():
    args = parse_args()

    excel_path = Path(args.excel_file) if args.excel_file else DEFAULT_EXCEL_PATH

    if not excel_path.exists():
        print(f"ERROR: Excel file not found at {excel_path}")
        sys.exit(1)

    configs = read_adjust_configs(excel_path)
    if not configs:
        print("No valid adjustment configs found in Excel file.")
        sys.exit(0)

    # Preview
    print(f"Loaded {len(configs)} order config(s) from {excel_path.name}:")
    print()
    for cfg in configs:
        tick_label = " [TICK_BASED]" if cfg.tick_based else ""
        print(f"  {cfg.action} CONID={cfg.conid}  QTY={cfg.qty:g}{tick_label}")
        print(f"    Tiers:")
        for i, tier in enumerate(cfg.tiers):
            delta = tier["delta"]
            interval_min = tier["interval_sec"] / 60
            threshold = tier["threshold"]
            if threshold is not None:
                if cfg.action == "BUY":
                    threshold_str = f"until price >= ${threshold:.2f}"
                else:
                    threshold_str = f"until price <= ${threshold:.2f}"
            else:
                threshold_str = "until filled"
            if cfg.tick_based:
                delta_str = f"{delta:.0f} ticks"
            else:
                delta_str = f"${delta:.2f}"
            print(f"      T{i+1}: {delta_str} every {interval_min:.2f}min {threshold_str}")
        print()
    print()

    # Create adjuster and connect
    adjuster = IBOrderAdjuster()

    print(f"[{datetime.now().strftime('%H:%M:%S')}] Connecting to IB on {DEFAULT_HOST}:{args.port}...")
    adjuster.connect(DEFAULT_HOST, args.port, args.client_id)

    # Start message processing thread
    api_thread = threading.Thread(target=adjuster.run)
    api_thread.daemon = True
    api_thread.start()

    # Wait for connection
    if not adjuster.connected_event.wait(timeout=10):
        print("Failed to connect to TWS/Gateway")
        sys.exit(1)

    try:
        # Fetch initial prices for each config
        print()
        print("Fetching market prices...")
        initial_prices = {}
        for cfg in configs:
            try:
                price = adjuster.get_initial_price(cfg.conid, cfg.action)
                initial_prices[cfg.conid] = price
                side_label = "bid" if cfg.action == "SELL" else "ask"
                print(f"  CONID={cfg.conid} {cfg.action}: {side_label} = ${price:.4f}")
            except RuntimeError as e:
                print(f"  ERROR: {e}")
                adjuster.disconnect()
                sys.exit(1)

        print()

        # Place orders with transmit=False
        print("Placing orders (untransmitted)...")
        for cfg in configs:
            price = initial_prices[cfg.conid]
            adjuster.place_order_for_config(cfg, price)
            time.sleep(0.1)  # Small delay between placements
            print(f"  Placed order #{cfg.order_id}: {cfg.action} {cfg.qty:g} @ ${price:.4f} "
                  f"(CONID={cfg.conid}, untransmitted)")

        print()
        print(">>> Transmit orders in TWS to begin adjusting <<<")
        print()

        # Wait for user to transmit all orders in TWS
        adjuster.all_transmitted_event.wait()

        # Fetch current prices (user may have changed them before transmitting)
        adjuster.request_open_orders()
        print()

        # Run the adjustment loop
        run_adjustment_loop(adjuster, configs)

    except Exception as e:
        print(f"[ERROR] {e}")
        raise
    finally:
        adjuster.running = False
        adjuster.disconnect()
        print(f"[{datetime.now().strftime('%H:%M:%S')}] Disconnected")


if __name__ == "__main__":
    main()
