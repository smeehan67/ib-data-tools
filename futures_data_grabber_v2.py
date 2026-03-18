"""
futures_data_grabber_v2.py — Batched-connection variant of futures_data_grabber.py

Instead of one IB connection per symbol, this version batches all requests
onto a single connection using unique request IDs. This eliminates repeated
connect/disconnect overhead.

Performance: Runs in less than 10% of the time of the original
futures_data_grabber.py (e.g. ~8 seconds vs ~95 seconds for 27 symbols
in the validation/update path).

Data store: The Excel workbook is the primary data store. Existing data is
read from the workbook for validation (not individual CSVs). Individual CSV
output is controlled by the INDIVIDUAL_CSVS flag (disabled by default).

Two-phase approach:
  Phase 1: One connection downloads 1 month of data for ALL symbols that have
           existing data in the workbook (validation). Symbols without data
           skip this.
  Phase 2: One connection downloads 3 years of data for any symbols that
           failed validation or had no existing data.
"""

from ibapi.client import *
from ibapi.wrapper import *
from ibapi.contract import Contract
import csv
import os
import subprocess
import sys
import threading
from datetime import datetime
from openpyxl import Workbook, load_workbook


# ================================================================
# CONFIGURATION
# ================================================================

SYMBOLS_FILE = "futures_historical_data.csv"
VALIDATION_ROWS = 10
MAX_DATA_ROWS = 503
COMBINED_FILE = ""                       # Set to "" to disable combined CSV output
EXCEL_FILE = "futures_combined.xlsx"     # Set to "" to disable Excel workbook output
INDIVIDUAL_CSVS = False                  # Set to True to write individual {SYMBOL}.csv files
EXCEL_ZOOM_PERCENT = 210                 # Zoom level when opening Excel
OPEN_EXCEL_IN_BACKGROUND = True          # Set to True to open Excel without stealing focus


# ================================================================
# UTILITY FUNCTIONS
# ================================================================

def read_symbols_csv(file_name):
    """Read symbols CSV and return list of (symbol, exchange, conid) tuples.
    conid is "" if not provided."""
    symbols = []
    with open(file_name, "r") as f:
        reader = csv.reader(f)
        next(reader)  # skip header
        for row in reader:
            symbol = row[0].strip()
            exchange = row[1].strip()
            conid = row[2].strip() if len(row) > 2 else ""
            if symbol:
                symbols.append((symbol, exchange, conid))
    return symbols


def make_contract(symbol, exchange, conid=""):
    contract = Contract()
    contract.symbol = symbol
    contract.exchange = exchange
    if conid:
        contract.conId = int(conid)
        contract.secType = "STK"
    else:
        contract.secType = "CONTFUT"
    return contract


def read_existing_csv(file_name):
    """Read existing CSV and return list of (date_str, close_str) tuples."""
    rows = []
    with open(file_name, "r") as f:
        reader = csv.reader(f)
        next(reader)  # skip header
        for row in reader:
            rows.append((row[0], row[1]))
    return rows


def format_close(close_val):
    """Normalize a close value to a string matching CSV format (strip trailing zeros)."""
    if isinstance(close_val, str):
        return close_val
    return f"{close_val:g}"


def read_existing_from_excel(wb, symbol):
    """Read existing data from an Excel workbook worksheet for the given symbol.

    Returns list of (date_str, close_str) tuples, or empty list if worksheet not found.
    """
    if symbol not in wb.sheetnames:
        return []
    ws = wb[symbol]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:  # skip header
            continue
        date_val, close_val = row[0], row[1]
        rows.append((str(date_val), format_close(close_val)))
    return rows


def close_excel_file(filepath):
    """Close file if open in Excel on macOS."""
    try:
        filename = os.path.basename(filepath)
        script = f'''
        tell application "Microsoft Excel"
            if it is running then
                set wbCount to count of workbooks
                repeat with i from 1 to wbCount
                    set w to workbook i
                    if name of w is "{filename}" then
                        close w saving no
                        exit repeat
                    end if
                end repeat
            end if
        end tell
        '''
        subprocess.run(['osascript', '-e', script], capture_output=True, timeout=5)
        print(f"Closed {filename} if it was open in Excel")
    except Exception as e:
        print(f"Note: Could not close Excel file: {e}")


def open_excel_file(filepath, zoom=210, background=False):
    """Open file in Excel and set zoom to 210%."""
    try:
        filepath = os.path.abspath(filepath)
        if background:
            script = f'''
tell application "System Events"
    set frontApp to name of first application process whose frontmost is true
end tell
set activeWB to ""
tell application "Microsoft Excel"
    if it is running then
        try
            set activeWB to name of active workbook
        end try
    end if
    open POSIX file "{filepath}"
end tell
delay 1
if frontApp is "Microsoft Excel" and activeWB is not "" then
    set wbMenuName to activeWB
    if wbMenuName ends with ".xlsx" or wbMenuName ends with ".xlsm" then
        set wbMenuName to text 1 thru -6 of wbMenuName
    else if wbMenuName ends with ".xls" then
        set wbMenuName to text 1 thru -5 of wbMenuName
    end if
    tell application "System Events"
        tell process "Microsoft Excel"
            click menu bar item "Window" of menu bar 1
            delay 0.3
            click menu item wbMenuName of menu 1 of menu bar item "Window" of menu bar 1
        end tell
    end tell
else
    tell application frontApp to activate
end if
'''
            subprocess.run(['osascript', '-e', script], capture_output=True, timeout=15)
            print(f"Opened {filepath} in Excel (background, focus restored)")
        else:
            subprocess.run(['open', filepath], check=True)
            if zoom != 100:
                import time
                time.sleep(2)  # Wait for Excel to open the file
                script = f'''
            tell application "Microsoft Excel"
                activate
                set zoom of active window to {zoom}
            end tell
            '''
                subprocess.run(['osascript', '-e', script], capture_output=True, timeout=5)
                print(f"Opened {filepath} in Excel at {zoom}% zoom")
            else:
                print(f"Opened {filepath} in Excel")
    except Exception as e:
        print(f"ERROR: Could not open file: {e}")


# ================================================================
# BATCH DOWNLOADER CLASS
# ================================================================

class BatchDownloader(EWrapper, EClient):
    """Downloads historical data for multiple symbols on a single IB connection.

    Usage:
        app = BatchDownloader()
        app.requests = {req_id: (symbol, exchange, conid, duration), ...}
        app.connect(...)
        app.run()
        # app.results[req_id] = [(date, close), ...]
    """

    def __init__(self):
        EClient.__init__(self, self)
        self.requests = {}      # req_id -> (symbol, exchange, conid, duration)
        self.results = {}       # req_id -> [(date, close), ...]
        self.completed = set()
        self.lock = threading.Lock()

    def error(self, reqId, errorCode, errorString, advancedOrderRejectJson=""):
        if errorCode not in (2104, 2106, 2158):
            print(f"Error: reqId={reqId} code={errorCode} msg={errorString}")

    def nextValidId(self, orderId):
        self.nextOrderId = orderId
        self.start_all()

    def historicalData(self, reqId, bar):
        symbol = self.requests[reqId][0] if reqId in self.requests else "?"
        print(f"[{symbol}] Date: {bar.date}  Close: {bar.close}")
        if reqId not in self.results:
            self.results[reqId] = []
        self.results[reqId].append((bar.date, bar.close))

    def historicalDataEnd(self, reqId, start, end):
        symbol = self.requests[reqId][0] if reqId in self.requests else "?"
        count = len(self.results.get(reqId, []))
        print(f"\n[{symbol}] Historical data complete. {count} bars received.")
        with self.lock:
            self.completed.add(reqId)
            if len(self.completed) >= len(self.requests):
                print(f"\nAll {len(self.requests)} requests complete.")
                threading.Timer(1, self.stop).start()

    def start_all(self):
        """Fire off all queued historical data requests."""
        for req_id, (symbol, exchange, conid, duration) in self.requests.items():
            what_to_show = "ADJUSTED_LAST" if conid else "TRADES"
            self.reqHistoricalData(
                reqId=req_id,
                contract=make_contract(symbol, exchange, conid),
                endDateTime="",
                durationStr=duration,
                barSizeSetting="1 day",
                whatToShow=what_to_show,
                useRTH=0,
                formatDate=1,
                keepUpToDate=False,
                chartOptions=[]
            )

    def stop(self):
        self.disconnect()


def run_batch(requests, client_id=0):
    """Run a batch of historical data requests on a single connection.

    Args:
        requests: dict of {req_id: (symbol, exchange, conid, duration)}
        client_id: IB client ID

    Returns:
        dict of {req_id: [(date, close), ...]}
    """
    if not requests:
        return {}

    app = BatchDownloader()
    app.requests = requests
    app.nextOrderId = 0
    app.connect("127.0.0.1", 7497, client_id)
    app.run()
    return app.results


# ================================================================
# VALIDATION AND FILE UPDATE LOGIC
# ================================================================

def validate_and_update(symbol, existing, api_data):
    """Compare existing data against API data and update if valid.

    Returns:
        (status, final_rows) where status is "updated", "up to date",
        "mismatch", or "insufficient_api_data".
        final_rows is the data to use (or None if full download needed).
    """
    if len(api_data) < VALIDATION_ROWS + 1:
        print(f"  [{symbol}] API returned fewer than {VALIDATION_ROWS + 1} bars.")
        return "insufficient_api_data", None

    # Exclude the last row from both (may be incomplete today bar)
    # Take the 10 rows just before the last row
    file_check = existing[-VALIDATION_ROWS - 1:-1]
    api_check = api_data[-VALIDATION_ROWS - 1:-1]

    # Compare: align by date
    api_dates = {row[0]: row[1] for row in api_check}
    matched = 0
    mismatched = 0

    for date_str, close_str in file_check:
        if date_str in api_dates:
            api_close = format_close(api_dates[date_str])
            file_close = format_close(float(close_str))
            if api_close == file_close:
                matched += 1
            else:
                mismatched += 1
                print(f"  [{symbol}] MISMATCH on {date_str}: file={close_str} api={api_close}")

    print(f"  [{symbol}] Validation: {matched} matched, {mismatched} mismatched "
          f"out of {matched + mismatched} overlapping rows.")

    if mismatched == 0 and matched > 0:
        # Extract new rows from the already-downloaded data
        # Include last date (>=) to replace potentially incomplete bar
        last_date_str = existing[-1][0]
        new_rows = [(d, c) for d, c in api_data if d >= last_date_str]

        if not new_rows:
            print(f"  [{symbol}] No new rows. File is already up to date.")
            return "up to date", existing

        # Keep all existing rows except the last, then add API rows
        all_rows = existing[:-1] + [(d, format_close(c)) for d, c in new_rows]
        all_rows = all_rows[-MAX_DATA_ROWS:]

        print(f"  [{symbol}] Updated last row and appended {len(new_rows) - 1} new rows.")
        return "updated", all_rows
    else:
        print(f"  [{symbol}] Data mismatch detected.")
        return "mismatch", None


def prepare_full_download(symbol, data):
    """Trim and format full download data. Returns final_rows."""
    data = data[-MAX_DATA_ROWS:]
    final_rows = [(d, format_close(c)) for d, c in data]
    print(f"  [{symbol}] Prepared {len(final_rows)} rows from full download.")
    return final_rows


# ================================================================
# MAIN
# ================================================================

def main():
    import time

    start_time = time.time()
    symbols_file = sys.argv[1] if len(sys.argv) > 1 else SYMBOLS_FILE
    excel_file = sys.argv[2] if len(sys.argv) > 2 else EXCEL_FILE
    base_client_id = int(sys.argv[3]) if len(sys.argv) > 3 else 25

    symbols = read_symbols_csv(symbols_file)
    print(f"Loaded {len(symbols)} symbols from {symbols_file}\n")

    results = {}        # symbol -> status string
    symbol_data = {}    # symbol -> list of (date, close) tuples
    existing_data = {}  # symbol -> existing rows (for symbols with data in workbook)

    # Load existing Excel workbook (if it exists) for validation
    existing_wb = None
    if excel_file and os.path.exists(excel_file):
        existing_wb = load_workbook(excel_file, read_only=True, data_only=True)
        print(f"Loaded existing workbook: {excel_file}\n")

    # Classify symbols: which ones have existing data and need validation,
    # and which ones need a full download from the start
    need_validation = []
    need_full = []

    for symbol, exchange, conid in symbols:
        if existing_wb is not None:
            existing = read_existing_from_excel(existing_wb, symbol)
        else:
            existing = []

        if existing:
            if len(existing) < VALIDATION_ROWS + 1:
                print(f"  [{symbol}] Workbook has fewer than {VALIDATION_ROWS + 1} rows. Will do full download.")
                need_full.append((symbol, exchange, conid))
            else:
                existing_data[symbol] = existing
                need_validation.append((symbol, exchange, conid))
        else:
            print(f"  [{symbol}] No existing data. Will do full download.")
            need_full.append((symbol, exchange, conid))

    if existing_wb is not None:
        existing_wb.close()

    # --------------------------------------------------------
    # PHASE 1: Batch validation (1 month each, single connection)
    # --------------------------------------------------------
    if need_validation:
        print(f"\n{'=' * 60}")
        print(f"  PHASE 1: Validating {len(need_validation)} symbols (1 connection)")
        print(f"{'=' * 60}\n")

        val_requests = {}
        val_req_map = {}  # req_id -> (symbol, exchange, conid)
        for i, (symbol, exchange, conid) in enumerate(need_validation):
            req_id = i
            val_requests[req_id] = (symbol, exchange, conid, "1 M")
            val_req_map[req_id] = (symbol, exchange, conid)

        val_results = run_batch(val_requests, client_id=base_client_id)  # Validation phase

        # Process validation results
        for req_id, (symbol, exchange, conid) in val_req_map.items():
            api_data = val_results.get(req_id, [])
            existing = existing_data[symbol]

            status, final_rows = validate_and_update(symbol, existing, api_data)

            if final_rows is not None:
                results[symbol] = status
                symbol_data[symbol] = final_rows
            else:
                # Need full download
                need_full.append((symbol, exchange, conid))

    # --------------------------------------------------------
    # PHASE 2: Batch full downloads (3 years each, single connection)
    # --------------------------------------------------------
    if need_full:
        print(f"\n{'=' * 60}")
        print(f"  PHASE 2: Full download for {len(need_full)} symbols (1 connection)")
        print(f"{'=' * 60}\n")

        full_requests = {}
        full_req_map = {}
        for i, (symbol, exchange, conid) in enumerate(need_full):
            req_id = i
            full_requests[req_id] = (symbol, exchange, conid, "3 Y")
            full_req_map[req_id] = symbol

        full_results = run_batch(full_requests, client_id=base_client_id + 1)  # Download phase

        for req_id, symbol in full_req_map.items():
            data = full_results.get(req_id, [])
            if not data:
                print(f"  [{symbol}] No data received.")
                results[symbol] = "no data"
                symbol_data[symbol] = []
            else:
                final_rows = prepare_full_download(symbol, data)
                results[symbol] = "full download"
                symbol_data[symbol] = final_rows

    # --------------------------------------------------------
    # INDIVIDUAL CSV OUTPUT FILES (optional)
    # --------------------------------------------------------
    if INDIVIDUAL_CSVS:
        for symbol, _exchange, _conid in symbols:
            rows = symbol_data.get(symbol, [])
            if rows:
                file_name = f"{symbol}.csv"
                with open(file_name, "w", newline="") as csvfile:
                    writer = csv.writer(csvfile)
                    writer.writerow(["Date", "Close"])
                    for date, close in rows:
                        writer.writerow([date, close])
                print(f"  [{symbol}] Saved {len(rows)} rows to {file_name}")

    # --------------------------------------------------------
    # COMBINED CSV OUTPUT FILE
    # --------------------------------------------------------
    if COMBINED_FILE:
        with open(COMBINED_FILE, "w", newline="") as csvfile:
            writer = csv.writer(csvfile)
            writer.writerow(["Symbol", "Date", "Close"])
            for symbol, _exchange, _conid in symbols:
                for date, close in symbol_data.get(symbol, []):
                    writer.writerow([symbol, date, close])
        print(f"\nSaved combined data to {COMBINED_FILE}")

    # --------------------------------------------------------
    # EXCEL WORKBOOK OUTPUT (one worksheet per symbol)
    # --------------------------------------------------------
    if excel_file:
        close_excel_file(excel_file)
        if os.path.exists(excel_file):
            wb = load_workbook(excel_file)
        else:
            wb = Workbook()
            wb.remove(wb.active)  # remove default empty sheet
        for symbol, _exchange, _conid in symbols:
            # Remove existing worksheet for this symbol, then recreate it
            if symbol in wb.sheetnames:
                wb.remove(wb[symbol])
            ws = wb.create_sheet(title=symbol)
            ws.append(["Date", "Close"])
            for date, close in symbol_data.get(symbol, []):
                ws.append([int(date), float(close)])
        wb.save(excel_file)
        print(f"Saved Excel workbook to {excel_file}")
        open_excel_file(excel_file, zoom=EXCEL_ZOOM_PERCENT, background=OPEN_EXCEL_IN_BACKGROUND)

    # --------------------------------------------------------
    # SUMMARY
    # --------------------------------------------------------
    elapsed = time.time() - start_time
    print(f"\n{'=' * 60}")
    print(f"  DOWNLOAD COMPLETE — {len(results)} symbols processed")
    print(f"  Elapsed time: {elapsed:.1f} seconds")
    print(f"{'=' * 60}")
    for symbol, _exchange, _conid in symbols:
        status = results.get(symbol, "skipped")
        print(f"  {symbol:8s} → {status}")

    os.system('say "Futures historical data download complete"')


if __name__ == "__main__":
    main()
